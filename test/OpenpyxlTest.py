import openpyxl

workbook = openpyxl.load_workbook('E:\\test.xlsx')

worksheet = workbook['Sheet1']

row3=[item.value for item in list(worksheet.rows)[1]]

print('第2行值',row3)

col3=[item.value for item in list(worksheet.columns)[1]]

print('第2列值',col3)

cell_2_3=worksheet.cell(row=1,column=1).value

print('第2行第3列值',cell_2_3)

max_row=worksheet.max_row

print('最大行',max_row)
